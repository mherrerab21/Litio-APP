import streamlit as st
import pandas as pd
import yfinance as yf
import plotly.express as px
from datetime import datetime, timedelta
import dash
from dash import dcc, html
from dash.dependencies import Input, Output
from openpyxl import load_workbook

# URL del logo de la compañía en GitHub
logo_url = 'https://github.com/mherrerab21/Litio-APP/raw/main/arrayan-logo.png'

# Mostrar el logo de la compañía en el dashboard
st.image(logo_url, width=200)  # Ajusta el ancho según sea necesario

# Establecer el título de la página
st.title("Dashboard de Precios de Contrato de Litio")
st.markdown("---")

# Cambiar el color de fondo de la página a un verde militar
st.markdown(
    """
    <style>
    body {
        background-color: #4B5320;
        color: white;
    }
    .css-1h2yj8v {
        max-width: none !important;
    }
    </style>
    """,
    unsafe_allow_html=True
)

# Establecer todos los títulos con el mismo tamaño
st.markdown(
    """
    <style>
    .css-1xjvdi2 {
        font-size: 28px !important;
        max-width: none !important;
    }
    </style>
    """,
    unsafe_allow_html=True
)

# Initialize the Dash app
app = dash.Dash(__name__)

# Define the layout of the dashboard
app.layout = html.Div([
    dcc.Graph(id='latest-graph'),
    html.Div([
        html.Label('Seleccione el conjunto de datos a mostrar:'),
        dcc.RadioItems(
            id='data-selector',
            options=[
                {'label': 'Contrato 2407', 'value': 'df_lc2407'},
                {'label': 'Market', 'value': 'df_market'}
            ],
            value='df_lc2407',
            labelStyle={'display': 'block'}
        )
    ], style={'margin-top': '20px'})
])

# Define callback to update the graph based on the selected dataset
@app.callback(
    Output('latest-graph', 'figure'),
    [Input('data-selector', 'value')]
)
def update_graph(selected_data):
    if selected_data == 'df_lc2407':
        # Plot 'Latest' column from df_lc2407
        fig = px.line(df_lc2407, x=df_lc2407.index, y='Latest', labels={'index': 'Date', 'Latest': 'Price'})
    else:
        # Plot 'Latest' column from df_market
        fig = px.line(df_market, x=df_market.index, y='Latest', labels={'index': 'Date', 'Latest': 'Price'})
    fig.update_layout(title='Latest Prices Over Time', xaxis_title='Date', yaxis_title='Price')
    return fig

# Download Yahoo Finance Data (improved error handling)
ticker = 'CNY=X'  # Double-check the ticker symbol
today = datetime.today()
start_date = today - timedelta(days=7)  # Start a week ago
end_date = today

try:
    data_yahoo = yf.download(ticker, start=start_date, end=end_date)
    if not data_yahoo.empty:  # Check if data is empty
        cierre_ajustado = data_yahoo['Close'].iloc[0]  # Access closing price
    else:
        st.warning("No data downloaded for {} in the specified date range.".format(ticker))
except Exception as e:  # Handle different exceptions
    st.error("Error downloading data for {}: {}".format(ticker, e))

# Read Data from Excel (df_market)
nombre_archivo_excel = 'futuros litio.xlsx'
df_market = pd.read_excel(nombre_archivo_excel, sheet_name='Market')

# Data Cleaning and Transformation (corrected renaming)
df_market.rename(columns={'Conversion en Notas': 'Fecha'}, inplace=True)
df_market.set_index('Fecha', inplace=True)

# Column Name Cleaning and Removal
# (Code for this section is skipped as it's repeated from the original code)

# Convertir el índice a objetos de fecha y luego formatear las fechas
df_market.index = pd.to_datetime(df_market.index).strftime('%d/%m/%y')

# Función para obtener tipo de cambio de USD a CNY desde Yahoo Finance
def obtener_tipo_cambio(ticker):
    try:
        data_yahoo = yf.download(ticker, start=start_date, end=end_date)
        if not data_yahoo.empty:  # Check if data is empty
            return round(data_yahoo['Close'].iloc[0], 2)  # Access closing price and round to 2 decimal places
        else:
            st.warning("No data downloaded for {} in the specified date range.".format(ticker))
            return None
    except Exception as e:  # Handle different exceptions
        st.error("Error downloading data for {}: {}".format(ticker, e))
        return None

# Obtener tipo de cambio de USD a CNY desde Yahoo Finance
tipo_cambio_USD_CNY = obtener_tipo_cambio('USDCNY=X')

# Definir factor de conversión de KG a metric tons
factor_conversion_KG_to_MT = 1000  # 1 KG = 0.001 metric tons

# Data Cleaning and Transformation for df_lc2407 (from the second code snippet)
nombre_archivo_excel_lc2407 = r'C:\Users\maxhe\Scrappers\Litio\futuros litio.xlsx'
book_lc2407 = load_workbook(nombre_archivo_excel_lc2407, data_only=True)
hoja_market_lc2407 = book_lc2407['Contrato 2407']
data_lc2407 = []
for row in hoja_market_lc2407.iter_rows(values_only=True):
    data_lc2407.append(row)

column_headers_lc2407 = data_lc2407[0]
df_lc2407 = pd.DataFrame(data_lc2407[1:], columns=column_headers_lc2407)
book_lc2407.close()

# Data Cleaning and Transformation (corrected renaming)
df_lc2407.set_index('Date', inplace=True)
df_lc2407.index = pd.to_datetime(df_lc2407.index).strftime('%d/%m/%y')

# Convertir las columnas 'Var%' y 'O.I%' a formato de porcentaje
df_lc2407['Var %'] = df_lc2407['Var %'] * 100
df_lc2407['O.I %'] = df_lc2407['O.I %'] * 100

# Mostrar DataFrame actualizado
st.markdown("## Precios de Contrato:")
st.dataframe(df_market)

# Mostrar tabla de variaciones junto con los precios de contrato
st.markdown("## Variaciones Porcentuales entre el último y penúltimo dato:")
st.dataframe(variacion_porcentual.to_frame(name='Variaciones Porcentuales'), height=400)

# Graficar los precios de contrato seleccionados a lo largo del tiempo
st.markdown("## Precios de Contrato a lo largo del Tiempo")
selected_prices = st.multiselect("Seleccionar Precio de Contrato:", df_market.columns)
if selected_prices:
    df_market_selected = df_market[selected_prices].reset_index()
    df_market_selected_long = pd.melt(df_market_selected, id_vars=['Fecha'], value_vars=selected_prices)
    fig = px.line(df_market_selected_long, x='Fecha', y='value', color='variable', labels={'Fecha': 'Fecha', 'value': 'Precio', 'variable': 'Precio de Contrato'})
    fig.update_layout(title="Precios de Contrato a lo largo del Tiempo", xaxis_title="Fecha", yaxis_title="Precio", legend_title="Precio de Contrato", width=1600, height=600)  # Ajusta el ancho del gráfico
    fig.update_traces(hovertemplate='%{x}<br>%{y}')
    st.plotly_chart(fig, use_container_width=False, config={'displayModeBar': True, 'scrollZoom': False})
else:
    st.warning("Por favor selecciona al menos un precio de contrato.")

if __name__ == '__main__':
    app.run_server(debug=True)
